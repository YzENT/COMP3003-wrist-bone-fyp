"""
Runs the saved model on inference_features.csv and produces a results spreadsheet.
Columns: Patient_ID, 1_Ulnar, 2_Radius, 3_Scaphoid, ... 10_Pisiform
A cell is marked 1 (bold red) if the model predicts fracture for that bone, else empty.

The output spreadsheet and text summary are saved alongside the inference CSV.

Usage: python generate_fracture_results.py <inference_features_csv>
"""

import sys
import glob
import pandas as pd
import joblib
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

MODEL_DIR = r"C:\\npy_128_128_48\\!!!!code submissions\\inference\\classification\\models"

BONE_LABELS = {
    1: 'ulnar',
    2: 'radius',
    3: 'scaphoid',
    4: 'lunate',
    5: 'triquetrum',
    6: 'hamate',
    7: 'trapezoid',
    8: 'capitate',
    9: 'trapezium',
    10: 'pisiform',
}

BONE_COLUMNS = [f"{idx}_{name.capitalize()}" for idx, name in BONE_LABELS.items()]


def load_model_and_scaler(model_dir):
    model_dir = Path(model_dir)
    model_files  = glob.glob(str(model_dir / "best_model_*.pkl"))
    scaler_files = glob.glob(str(model_dir / "best_scaler_*.pkl"))

    if not model_files:
        raise FileNotFoundError(f"No best_model_*.pkl found in {model_dir}")
    if not scaler_files:
        raise FileNotFoundError(f"No best_scaler_*.pkl found in {model_dir}")

    model_path  = model_files[0]
    scaler_path = scaler_files[0]

    print(f"Loading model : {model_path}")
    print(f"Loading scaler: {scaler_path}")
    return joblib.load(model_path), joblib.load(scaler_path)


def extract_patient_id(dataset_id: str) -> str:
    # UX001_scaphoid -> UX001
    return "_".join(dataset_id.split("_")[:-1])


def extract_bone_name(dataset_id: str) -> str:
    # UX001_scaphoid -> scaphoid
    return dataset_id.split("_")[-1].lower()


def predict_fracture_spreadsheet(inference_csv):
    inference_csv = Path(inference_csv)
    model, scaler = load_model_and_scaler(MODEL_DIR)

    df = pd.read_csv(inference_csv)


    # model = joblib.load('best_model_svm_rbf_minmax.pkl')
    # scaler = joblib.load('best_scaler_minmax.pkl')
    # y_pred = model.predict(scaler.transform(X_new))

    feature_cols = [c for c in df.columns if c != 'dataset_id']
    X_scaled = scaler.transform(df[feature_cols].values)
    predictions = model.predict(X_scaled) # 0 = healthy, 1 = fracture

    results = {}
    for dataset_id, pred in zip(df['dataset_id'], predictions):
        patient_id = extract_patient_id(dataset_id)
        bone_name = extract_bone_name(dataset_id)
        results.setdefault(patient_id, {})[bone_name] = int(pred)

    sorted_patients = sorted(results.keys())

    # spreadsheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Fracture Results"

    font = Font(name="Calibri", size=11)
    header_font = Font(name="Calibri", size=11, bold=True)
    fracture_font = Font(name="Calibri", size=11, bold=True, color="FF0000")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left",   vertical="center")

    headers = ["Patient_ID"] + BONE_COLUMNS
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_align

    ws.freeze_panes = "A2" # freeze header
    ws.row_dimensions[1].height = 15.75
    ws.column_dimensions["A"].width = 35.57
    for col_idx in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14.86

    for row_idx, patient_id in enumerate(sorted_patients, start=2):
        patient_bones = results[patient_id]

        id_cell = ws.cell(row=row_idx, column=1, value=patient_id)
        id_cell.font = font
        id_cell.alignment = left_align

        for col_idx, bone_col in enumerate(BONE_COLUMNS, start=2):
            bone_name = bone_col.split("_", 1)[1].lower()
            pred = patient_bones.get(bone_name, None)

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = center_align

            if pred == 1:
                cell.value = 1
                cell.font = fracture_font
            else:
                cell.value = None
                cell.font = font

    output_path = inference_csv.parent / "fracture_results.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")

    # export as txt to double check spreadsheet code is correct
    # txt_path = inference_csv.parent / "fracture_predictions.txt"
    # with open(txt_path, "w") as f:
    #     f.write("FRACTURE PREDICTIONS SUMMARY\n")
    #     f.write("=" * 40 + "\n\n")
    #     total_fractures = 0
    #     for patient_id in sorted_patients:
    #         fractured_bones = [
    #             bone for bone, pred in results[patient_id].items() if pred == 1
    #         ]
    #         if fractured_bones:
    #             f.write(f"{patient_id}:\n")
    #             for bone in sorted(fractured_bones):
    #                 f.write(f"  - {bone}\n")
    #             f.write("\n")
    #             total_fractures += len(fractured_bones)
    #     f.write("=" * 40 + "\n")
    #     f.write(f"Total patients with fractures : {sum(1 for p in sorted_patients if any(v == 1 for v in results[p].values()))}/{len(sorted_patients)}\n")
    #     f.write(f"Total fractured bones : {total_fractures}\n")
    # print(f"Saved: {txt_path}")


if __name__ == "__main__":

    args = sys.argv[1:]

    if len(args) != 1:
        print("Usage: python generate_fracture_results.py <inference_features_csv>")
        sys.exit(1)

    input_csv = args[0]

    predict_fracture_spreadsheet(input_csv)